import requests
import time
import os
import openpyxl as xl 
from tqdm import tqdm 
from dotenv import load_dotenv

# Load the Excel workbook and select the sheet
wb = xl.load_workbook('csfail_database_reverse.xlsx')
sheet = wb['Arkusz1']

load_dotenv()

API = os.getenv("API")

# Initialize row count and game range
rows_count = sheet.max_row
start_game = 2750000
end_game = 5676636 - rows_count

# Loop through games in reverse order and fetch data
for i in tqdm(range(end_game, start_game, -1)):
    time.sleep(1)
    api = requests.get(f'{API}{i}')
    api = api.json()

    rows_count += 1
    items = ["crashedAt", "totalBankUsd", "usersCount", "itemsCount"]

    # Write the fetched data into the corresponding Excel columns
    for item in items:
        value = api["data"]["game"][item]
        sheet.cell(row=rows_count, column=items.index(item) + 1, value=value)

    wb.save('csfail_database_reverse.xlsx')
